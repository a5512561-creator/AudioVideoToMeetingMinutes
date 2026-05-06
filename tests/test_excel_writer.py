from openpyxl import load_workbook
from script.schemas import (
    Conclusion,
    Action,
    KeyPoint,
    MeetingMinutes,
    ReviewNote,
    ReviewResult,
)
from script.excel_writer import write_minutes_xlsx


INFERRED_FILL_HEX = "FFFFF2CC"  # light yellow


def _conc(text="C-text", inferred=False, speaker=None):
    return Conclusion(
        text=text, is_inferred=inferred, source_quote="q",
        source_timestamp="00:00:01", source_speaker=speaker,
    )


def _act(task="A-task", owner="o", inferred=False, speaker=None):
    return Action(
        task=task, owner=owner, due="2026-05-15", priority="high",
        source_quote="q", source_timestamp="00:00:02", source_speaker=speaker,
        rationale="r", is_inferred=inferred, owner_inferred=False,
        due_inferred=False, priority_inferred=True,
    )


def _kp(text="K-text", inferred=False, speaker=None):
    return KeyPoint(
        text=text, is_inferred=inferred, source_quote="kq",
        source_timestamp="00:00:03", source_speaker=speaker,
    )


def _ok_note(section, tid):
    return ReviewNote(target_section=section, target_id=tid,
                      category="ok", severity="info", note="", suggestion="")


def _warn_note(section, tid, note="msg"):
    return ReviewNote(target_section=section, target_id=tid,
                      category="ambiguity", severity="warn", note=note, suggestion="fix")


def test_xlsx_has_two_sheets(tmp_path):
    minutes = MeetingMinutes(conclusions=[_conc()], actions=[_act()])
    review = ReviewResult(notes=[_ok_note("conclusion", "C1"), _ok_note("action", "A1")])
    dst = tmp_path / "minutes.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    wb = load_workbook(dst)
    assert wb.sheetnames == ["會議記錄", "Review摘要"]


def test_inferred_conclusion_gets_prefix_and_yellow_fill(tmp_path):
    minutes = MeetingMinutes(conclusions=[_conc(text="預算 500 萬", inferred=True)], actions=[])
    review = ReviewResult(notes=[_ok_note("conclusion", "C1")])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    wb = load_workbook(dst)
    ws = wb["會議記錄"]
    found = False
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("[LLM推論]"):
                assert "預算 500 萬" in c.value
                assert c.fill.fgColor.rgb == INFERRED_FILL_HEX
                found = True
    assert found, "expected at least one [LLM推論]-prefixed yellow cell"


def test_review_column_filled_per_row(tmp_path):
    minutes = MeetingMinutes(
        conclusions=[_conc(text="c1"), _conc(text="c2")],
        actions=[_act(task="a1")],
    )
    review = ReviewResult(notes=[
        _ok_note("conclusion", "C1"),
        _warn_note("conclusion", "C2", note="模糊"),
        _ok_note("action", "A1"),
    ])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    ws = load_workbook(dst)["會議記錄"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("✅" in s for s in all_strs)
    assert any("⚠️" in s and "模糊" in s for s in all_strs)


def test_review_summary_sheet_contains_only_warn_or_error(tmp_path):
    minutes = MeetingMinutes(
        conclusions=[_conc(text="c1"), _conc(text="c2")],
        actions=[],
    )
    review = ReviewResult(notes=[
        _ok_note("conclusion", "C1"),
        _warn_note("conclusion", "C2", note="模糊"),
    ])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    ws = load_workbook(dst)["Review摘要"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("C2" in s for s in all_strs)
    assert not any(isinstance(s, str) and "C1" in s for s in all_strs)


def test_key_points_section_renders(tmp_path):
    minutes = MeetingMinutes(
        conclusions=[_conc(text="c1")],
        key_points=[_kp(text="spec 說明")],
        actions=[_act(task="a1")],
    )
    review = ReviewResult(notes=[
        _ok_note("conclusion", "C1"),
        _ok_note("key_point", "K1"),
        _ok_note("action", "A1"),
    ])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    wb = load_workbook(dst)
    ws = wb["會議記錄"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    # Section header exists
    assert "會議重點" in all_strs
    # K1 row text present
    assert "K1" in all_strs
    assert "spec 說明" in all_strs


def test_key_point_warn_appears_in_review_summary(tmp_path):
    minutes = MeetingMinutes(
        conclusions=[],
        key_points=[_kp(text="重要技術說明")],
        actions=[],
    )
    review = ReviewResult(notes=[
        _warn_note("key_point", "K1", note="說明不清楚"),
    ])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst))
    ws = load_workbook(dst)["Review摘要"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "重點" in all_strs  # sec_label for key_point
    assert "K1" in all_strs


def test_speaker_map_substitutes_labels(tmp_path):
    # conclusion + action with SPEAKER_00 / SPEAKER_01
    minutes = MeetingMinutes(
        conclusions=[_conc(text="c1", speaker="SPEAKER_00")],
        actions=[_act(task="a1", speaker="SPEAKER_01")],
    )
    review = ReviewResult(notes=[
        _ok_note("conclusion", "C1"),
        _ok_note("action", "A1"),
    ])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst),
                       speaker_map={"SPEAKER_00": "Albert", "SPEAKER_01": "John"})
    ws = load_workbook(dst)["會議記錄"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Albert" in s for s in all_strs)
    assert any("John" in s for s in all_strs)
    assert not any("SPEAKER_00" in s for s in all_strs)


def test_speaker_map_passthrough_when_no_mapping(tmp_path):
    minutes = MeetingMinutes(conclusions=[_conc(speaker="SPEAKER_99")], actions=[])
    review = ReviewResult(notes=[_ok_note("conclusion", "C1")])
    dst = tmp_path / "m.xlsx"
    write_minutes_xlsx(minutes, review, str(dst),
                       speaker_map={"SPEAKER_00": "Albert"})
    ws = load_workbook(dst)["會議記錄"]
    all_strs = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("SPEAKER_99" in s for s in all_strs)
