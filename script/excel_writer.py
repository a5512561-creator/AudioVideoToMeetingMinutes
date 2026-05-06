from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from script.schemas import (
    MeetingMinutes,
    ReviewResult,
    ReviewNote,
)


_INFERRED_PREFIX = "[LLM推論] "
_INFERRED_FILL = PatternFill("solid", fgColor="FFFFF2CC")
_HEADER_FILL = PatternFill("solid", fgColor="FF305496")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_COL_HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
_COL_HEADER_FONT = Font(bold=True)

_SEV_ICON = {"info": "✅", "warn": "⚠️", "error": "❌"}


def _prefix(text: str, inferred: bool) -> str:
    return f"{_INFERRED_PREFIX}{text}" if inferred else text


def _format_review(note: ReviewNote | None) -> str:
    if note is None:
        return ""
    icon = _SEV_ICON.get(note.severity, "")
    if note.category == "ok":
        return f"{icon} OK"
    return f"{icon} {note.category}：{note.note}"


def _index_review(review: ReviewResult) -> dict[tuple[str, str], ReviewNote]:
    return {(n.target_section, n.target_id): n for n in review.notes}


def write_minutes_xlsx(minutes: MeetingMinutes, review: ReviewResult, dst: str) -> None:
    Path(dst).parent.mkdir(parents=True, exist_ok=True)
    review_ix = _index_review(review)

    wb = Workbook()
    ws = wb.active
    ws.title = "會議記錄"

    # ---- Section: 會議結論 ----
    conc_headers = ["編號", "結論內容", "來源原句", "時間戳", "發言者", "Review檢查結果"]
    _write_section_header(ws, row=1, title="會議結論", span=len(conc_headers))
    _write_column_headers(ws, row=2, headers=conc_headers)
    next_row = 3
    for i, c in enumerate(minutes.conclusions, start=1):
        cid = f"C{i}"
        rev = review_ix.get(("conclusion", cid))
        text_cell = _prefix(c.text, c.is_inferred)
        ws.cell(row=next_row, column=1, value=cid)
        cell_text = ws.cell(row=next_row, column=2, value=text_cell)
        if c.is_inferred:
            cell_text.fill = _INFERRED_FILL
        ws.cell(row=next_row, column=3, value=c.source_quote)
        ws.cell(row=next_row, column=4, value=c.source_timestamp)
        ws.cell(row=next_row, column=5, value=c.source_speaker or "")
        ws.cell(row=next_row, column=6, value=_format_review(rev))
        next_row += 1

    # blank separator row
    next_row += 1

    # ---- Section: 會議重點 ----
    kp_headers = ["編號", "重點內容", "來源原句", "時間戳", "發言者", "Review檢查結果"]
    _write_section_header(ws, row=next_row, title="會議重點", span=len(kp_headers))
    _write_column_headers(ws, row=next_row + 1, headers=kp_headers)
    next_row += 2
    for i, k in enumerate(minutes.key_points, start=1):
        kid = f"K{i}"
        rev = review_ix.get(("key_point", kid))
        text_cell = _prefix(k.text, k.is_inferred)
        ws.cell(row=next_row, column=1, value=kid)
        cell_text = ws.cell(row=next_row, column=2, value=text_cell)
        if k.is_inferred:
            cell_text.fill = _INFERRED_FILL
        ws.cell(row=next_row, column=3, value=k.source_quote)
        ws.cell(row=next_row, column=4, value=k.source_timestamp)
        ws.cell(row=next_row, column=5, value=k.source_speaker or "")
        ws.cell(row=next_row, column=6, value=_format_review(rev))
        next_row += 1

    # blank separator row
    next_row += 1

    # ---- Section: Action Items ----
    act_headers = [
        "編號", "任務", "負責人", "期限", "優先度",
        "來源原句", "時間戳", "發言者", "LLM判斷依據", "Review檢查結果",
    ]
    _write_section_header(ws, row=next_row, title="Action Items", span=len(act_headers))
    _write_column_headers(ws, row=next_row + 1, headers=act_headers)
    next_row += 2
    for i, a in enumerate(minutes.actions, start=1):
        aid = f"A{i}"
        rev = review_ix.get(("action", aid))
        ws.cell(row=next_row, column=1, value=aid)

        task_cell = ws.cell(row=next_row, column=2, value=_prefix(a.task, a.is_inferred))
        if a.is_inferred:
            task_cell.fill = _INFERRED_FILL

        owner_cell = ws.cell(row=next_row, column=3, value=_prefix(a.owner, a.owner_inferred))
        if a.owner_inferred:
            owner_cell.fill = _INFERRED_FILL

        due_cell = ws.cell(row=next_row, column=4, value=_prefix(a.due, a.due_inferred))
        if a.due_inferred:
            due_cell.fill = _INFERRED_FILL

        prio_cell = ws.cell(
            row=next_row, column=5, value=_prefix(a.priority, a.priority_inferred)
        )
        if a.priority_inferred:
            prio_cell.fill = _INFERRED_FILL

        ws.cell(row=next_row, column=6, value=a.source_quote)
        ws.cell(row=next_row, column=7, value=a.source_timestamp)
        ws.cell(row=next_row, column=8, value=a.source_speaker or "")
        ws.cell(row=next_row, column=9, value=a.rationale)
        ws.cell(row=next_row, column=10, value=_format_review(rev))
        next_row += 1

    # column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A3"

    # ---- Sheet 2: Review 摘要 ----
    ws2 = wb.create_sheet("Review摘要")
    sum_headers = ["對應位置", "編號", "類型", "嚴重度", "說明", "建議修正"]
    _write_column_headers(ws2, row=1, headers=sum_headers)
    sec_label = {"conclusion": "結論", "key_point": "重點", "action": "Action"}
    r = 2
    for n in review.notes:
        if n.severity in ("warn", "error"):
            ws2.cell(row=r, column=1, value=sec_label[n.target_section])
            ws2.cell(row=r, column=2, value=n.target_id)
            ws2.cell(row=r, column=3, value=n.category)
            ws2.cell(row=r, column=4, value=n.severity)
            ws2.cell(row=r, column=5, value=n.note)
            ws2.cell(row=r, column=6, value=n.suggestion)
            r += 1
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    wb.save(dst)


def _write_section_header(ws, *, row: int, title: str, span: int) -> None:
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_column_headers(ws, *, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _COL_HEADER_FONT
        c.fill = _COL_HEADER_FILL
